import akshare as ak
import pandas as pd
import numpy as np
from ta.trend import MACD, ADXIndicator
from ta.momentum import StochasticOscillator, RSIIndicator
from ta.volume import MFIIndicator, OnBalanceVolumeIndicator, ChaikinMoneyFlowIndicator
from ta.volatility import BollingerBands
from datetime import datetime, timedelta
import os
import time
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import concurrent.futures
import random

# --- 1. 环境与配置 ---
current_dir = os.getcwd()
sys.path.append(current_dir)

CONFIG = {
    "MIN_AMOUNT": 20000000,   # 最低成交额 2000万 (过滤僵尸股)
    "MIN_PRICE": 2.5,         # 最低股价 (过滤垃圾股)
    "MAX_WORKERS": 12,        # 线程数 (安全并发)
    "DAYS_LOOKBACK": 150      # 数据回溯天数
}

HOT_CONCEPTS = [] 
HISTORY_FILE = "history_log.csv"

# --- 2. 基础数据获取 ---
def get_market_hot_spots():
    global HOT_CONCEPTS
    try:
        df = ak.stock_board_concept_name_em()
        df = df.sort_values(by="涨跌幅", ascending=False).head(15)
        HOT_CONCEPTS = df["板块名称"].tolist()
        print(f"🔥 今日风口: {HOT_CONCEPTS}")
    except:
        HOT_CONCEPTS = []
        print("⚠️ 热点获取失败，跳过热点匹配")

def get_targets_robust():
    print(">>> [1/4] 获取全市场股票并预过滤...")
    try:
        df = ak.stock_zh_a_spot_em()
        # 兼容性重命名
        col_map = {"最新价": "price", "最新价格": "price", "成交额": "amount", "成交金额": "amount", "代码": "code", "名称": "name"}
        df.rename(columns=col_map, inplace=True)
        
        # 强制转数字
        df["price"] = pd.to_numeric(df["price"], errors='coerce')
        df["amount"] = pd.to_numeric(df["amount"], errors='coerce')
        df.dropna(subset=["price", "amount"], inplace=True)
        
        # 基础过滤
        df = df[df["code"].str.startswith(("60", "00"))]
        df = df[~df['name'].str.contains('ST|退')]
        df = df[df["price"] >= CONFIG["MIN_PRICE"]]
        df = df[df["amount"] > CONFIG["MIN_AMOUNT"]]
        
        targets = df[["code", "name"]]
        print(f"✅ 有效标的: {len(targets)} 只")
        return targets, "在线API"
    except Exception as e:
        print(f"⚠️ 异常: {e}")
        return pd.DataFrame(), "无结果"

def get_data_with_retry(code, start_date):
    time.sleep(random.uniform(0.01, 0.05)) # 随机延迟防封
    for _ in range(2):
        try:
            df = ak.stock_zh_a_hist(symbol=code, period="daily", start_date=start_date, adjust="qfq", timeout=5)
            if df is None or df.empty: raise ValueError("Empty")
            return df
        except: time.sleep(0.2)
    return None

def get_60m_data(code):
    """获取60分钟K线数据 (分钟级接口)"""
    try:
        df = ak.stock_zh_a_hist_min_em(symbol=code, period="60", adjust="qfq")
        if df is None or df.empty: return None
        return df.tail(100)
    except: return None

def get_stock_catalysts(code):
    try:
        news_df = ak.stock_news_em(symbol=code)
        if not news_df.empty:
            return news_df.iloc[0]['新闻标题']
    except: pass
    return ""

# --- 3. 🔥 K线形态深度分析 (去伪存真) ---
def analyze_kline_health(df_full):
    """
    深度分析K线逻辑，区分洗盘与出货
    返回: (状态描述, 评分加成)
    """
    if len(df_full) < 60: return "⚪数据不足", 0
    
    curr = df_full.iloc[-1]
    
    # 基础计算
    body_top = max(curr['open'], curr['close'])
    body_bottom = min(curr['open'], curr['close'])
    price_range = curr['high'] - curr['low']
    
    if price_range == 0: return "⚪极小波动", 0
    
    upper_len = curr['high'] - body_top
    lower_len = body_bottom - curr['low']
    upper_ratio = upper_len / price_range
    lower_ratio = lower_len / price_range
    
    # 环境判定
    # 相对位置 (Rank): 0=低位, 1=高位
    high_60 = df_full['high'].tail(60).max()
    low_60 = df_full['low'].tail(60).min()
    rp = (curr['close'] - low_60) / (high_60 - low_60 + 0.0001)
    
    # 量比
    vol_ma5 = df_full['volume'].tail(5).mean()
    vol_ratio = curr['volume'] / vol_ma5 if vol_ma5 > 0 else 1.0
    
    # 趋势
    ma20_curr = df_full['close'].tail(20).mean()
    trend_up = curr['close'] > ma20_curr

    # --- 判定逻辑 ---
    # A. 长上影线 (占比>40%)
    if upper_ratio > 0.4:
        # 高位+放量 = 抛压
        if rp > 0.8 and vol_ratio > 2.0: return "⚠️高位抛压", -30
        # 下跌趋势+阴线 = 受阻
        elif not trend_up and curr['close'] < curr['open']: return "📉冲高受阻", -10
        # 低位+缩量/温和+阳线 = 仙人指路
        elif rp < 0.6 and vol_ratio < 1.5 and curr['close'] >= curr['open']: return "☝️仙人指路", 15
        else: return "⚪上影震荡", 0

    # B. 长下影线 (占比>40%)
    elif lower_ratio > 0.4:
        # 下降趋势+重心下移 = 中继
        if not trend_up and curr['close'] < df_full['close'].iloc[-2]: return "⚠️下跌中继", -20
        # 回踩支撑+缩量 = 金针探底
        elif curr['low'] <= ma20_curr and curr['close'] > ma20_curr: return "🛡️金针探底", 20
        # 低位 = 承接
        elif rp < 0.2: return "⚓底部承接", 15
        else: return "⚪下影震荡", 5

    # C. 实体阳线 (>60%)
    elif (curr['close'] - curr['open']) / price_range > 0.6:
        # 反包昨日阴线
        prev_open = df_full['open'].iloc[-2]
        prev_close = df_full['close'].iloc[-2]
        if prev_close < prev_open and curr['close'] > prev_open: return "⚡阳包阴", 25
        return "💪实体强攻", 10

    # D. 实体阴线
    elif (curr['open'] - curr['close']) / price_range > 0.6:
        if vol_ratio > 2.0: return "💚放量杀跌", -20
        return "🤢阴线调整", -5

    # E. 小星线
    else:
        if vol_ratio < 0.6: return "✨缩量十字", 5
            
    return "⚪普通震荡", 0

# --- 4. 核心处理逻辑 ---
def process_stock_logic(df, code, name):
    if len(df) < 100: return None
    
    # 清洗数据
    rename_dict = {"日期":"date","开盘":"open","收盘":"close","最高":"high","最低":"low","成交量":"volume","成交额":"amount"}
    col_map = {k:v for k,v in rename_dict.items() if k in df.columns}
    df.rename(columns=col_map, inplace=True)
    
    close = df["close"]
    volume = df["volume"]
    df["vwap"] = df["amount"] / volume if "amount" in df.columns else (df['high'] + df['low'] + close) / 3

    # 基础指标
    df["pct_chg"] = close.pct_change() * 100
    today_pct = df["pct_chg"].iloc[-1]
    pct_3day = (close.iloc[-1] - close.iloc[-4]) / close.iloc[-4] * 100 if len(close) > 4 else 0
    
    df["MA5"] = close.rolling(5).mean()
    df["MA20"] = close.rolling(20).mean()
    df["MA60"] = close.rolling(60).mean()
    df["BIAS20"] = (close - df["MA20"]) / df["MA20"] * 100

    bb_ind = BollingerBands(close, window=20, window_dev=2)
    df["BB_Upper"] = bb_ind.bollinger_hband()
    df["BB_Lower"] = bb_ind.bollinger_lband()
    df["BB_Width"] = bb_ind.bollinger_wband()
    df["BB_PctB"] = bb_ind.bollinger_pband()

    macd = MACD(close)
    df["DIF"] = macd.macd()
    df["DEA"] = macd.macd_signal()
    df["MACD_Bar"] = macd.macd_diff()
    
    kdj = StochasticOscillator(df['high'], df['low'], close)
    df["K"] = kdj.stoch()
    df["J"] = kdj.stoch() * 3 - kdj.stoch_signal() * 2
    
    df["RSI"] = RSIIndicator(close, window=14).rsi()
    
    obv_ind = OnBalanceVolumeIndicator(close, volume)
    df["OBV"] = obv_ind.on_balance_volume()
    df["OBV_MA10"] = df["OBV"].rolling(10).mean()
    
    cmf_ind = ChaikinMoneyFlowIndicator(df['high'], df['low'], close, volume, window=20)
    df["CMF"] = cmf_ind.chaikin_money_flow()
    df["ADX"] = ADXIndicator(df['high'], df['low'], close, window=14).adx()

    curr = df.iloc[-1]
    prev = df.iloc[-2]
    prev_2 = df.iloc[-3]

    # ================================
    # 🔥 1. 必杀熔断 (Fail Fast)
    # ================================
    if curr["J"] > 100: return None # 防追高
    if curr["OBV"] <= curr["OBV_MA10"]: return None # 资金流出必杀
    if curr["CMF"] < 0.05: return None # 资金强度不够
    if curr["CMF"] <= prev["CMF"]: return None # 资金必须加速流入 (今日>昨日)
    if curr["MACD_Bar"] <= prev["MACD_Bar"]: return None # 动能必须增强 (红增或绿缩)

    # ================================
    # 2. 策略判定
    # ================================
    signal_type = ""
    suggest_buy = curr["close"]
    stop_loss = curr["MA20"]
    has_zt = (df["pct_chg"].tail(30) > 9.5).sum() >= 1
    
    # 策略A: ⚱️ 黄金坑 (深跌反转)
    is_deep_dip = (prev["BIAS20"] < -8) or (prev["RSI"] < 25)
    is_reversal = (curr["close"] > curr["MA5"]) and (curr["pct_chg"] > 1.5)
    if is_deep_dip and is_reversal:
        signal_type = "⚱️黄金坑(企稳)"; stop_loss = round(curr["low"] * 0.98, 2)
    
    # 策略B: 🐉 龙回头 (强势股回调)
    if not signal_type and has_zt and curr["close"] > curr["MA60"]:
        if curr["volume"] < df["volume"].tail(30).max() * 0.6: # 缩量
            if -5.0 < curr["BIAS20"] < 8.0:
                signal_type = "🐉龙回头"; stop_loss = round(curr["BB_Lower"], 2)
    
    # 策略C: 🏦 机构控盘 (强趋势)
    if not signal_type and curr["close"] > curr["MA60"] and curr["CMF"] > 0.1 and curr["ADX"] > 25:
        signal_type = "🏦机构控盘"; suggest_buy = round(curr["vwap"], 2)
    
    # 策略D: ⚡ 底部变盘 (布林收口)
    if not signal_type and curr["close"] < curr["MA60"] * 1.2 and curr["BB_Width"] < 12:
        signal_type = "⚡底部变盘"

    # ================================
    # 3. 共振与形态
    # ================================
    # 筹码
    chip_signal = ""
    high_120 = df["high"].tail(120).max()
    low_120 = df["low"].tail(120).min()
    current_pos = (curr["close"] - low_120) / (high_120 - low_120 + 0.001)
    if current_pos < 0.4:
        volatility = df["close"].tail(60).std() / df["close"].tail(60).mean()
        if volatility < 0.15: chip_signal = "🏆筹码密集" 

    # 形态
    patterns = []
    # 红肥绿瘦
    vol_up = df[df['close']>df['open']].tail(20)['volume'].sum()
    vol_down = df[df['close']<df['open']].tail(20)['volume'].sum()
    if vol_up > vol_down * 2.0 and curr["close"] > curr["MA20"]: patterns.append("🟥红肥绿瘦")
    # N字反包
    if (prev['close'] < prev['open']) and (curr['close'] > curr['open']) and (curr['close'] > prev['open']): patterns.append("⚡N字反包")
    # 蚂蚁上树
    recent_5 = df.tail(5)
    if (recent_5['close'] > recent_5['MA5']).all() and (recent_5['pct_chg'].abs() < 4.0).all() and (recent_5['close'].iloc[-1] > recent_5['close'].iloc[0]):
        patterns.append("🐜蚂蚁上树")
    pattern_str = " ".join(patterns)

    # 金叉
    is_macd_gold = (prev["DIF"] < prev["DEA"]) and (curr["DIF"] > curr["DEA"])
    is_kdj_gold = (prev["J"] < prev["K"]) and (curr["J"] > curr["K"]) and (curr["J"] < 80)
    
    if signal_type != "⚱️黄金坑(企稳)":
        if not (is_macd_gold or is_kdj_gold): return None # 非黄金坑策略必须有金叉

    # --- 最终入围检查 ---
    has_strategy = bool(signal_type)
    has_resonance = bool(chip_signal and pattern_str) 
    if not (has_strategy or has_resonance): return None

    # 🔥 K线健康度分析
    kline_status, kline_score = analyze_kline_health(df)

    # ================================
    # 4. 60分钟择时
    # ================================
    status_60m = "⏳数据不足"
    try:
        df_60 = get_60m_data(code)
        if df_60 is not None and len(df_60) > 30:
            df_60.rename(columns={"时间":"date","开盘":"open","收盘":"close","最高":"high","最低":"low","成交量":"volume"}, inplace=True)
            close_60 = df_60["close"]
            macd_60 = MACD(close_60)
            dif_60, dea_60 = macd_60.macd(), macd_60.macd_signal()
            ma20_60 = close_60.rolling(20).mean()
            
            c60, ma20_curr = close_60.iloc[-1], ma20_60.iloc[-1]
            dif_curr, dea_curr = dif_60.iloc[-1], dea_60.iloc[-1]
            dif_prev, dea_prev = dif_60.iloc[-2], dea_60.iloc[-2]
            
            is_gold_60 = (dif_prev < dea_prev) and (dif_curr > dea_curr)
            if is_gold_60: status_60m = "✅60分金叉"
            elif dif_curr > dea_curr and c60 > ma20_curr: status_60m = "🚀60分多头"
            elif dif_curr < dea_curr: status_60m = "⚠️60分回调"
            else: status_60m = "⚪60分震荡"
    except: status_60m = "❌获取失败"

    # --- 组装 ---
    cross_status = ""
    if is_macd_gold and is_kdj_gold: cross_status = "⚡双金叉"
    elif is_macd_gold: cross_status = "🔥MACD金叉"
    elif is_kdj_gold: cross_status = "📈KDJ金叉"
    elif signal_type == "⚱️黄金坑(企稳)": cross_status = "🟢绿柱缩短"

    reasons = []
    if signal_type: reasons.append("策略")
    if has_resonance: reasons.append("筹/形共振")
    if cross_status == "⚡双金叉": reasons.append("双金叉")
    resonance_str = "+".join(reasons)

    news_title = get_stock_catalysts(code)
    hot_matched = ""
    for hot in HOT_CONCEPTS:
        if hot in news_title: hot_matched = hot; break
    display_concept = f"🔥{hot_matched}" if hot_matched else ""

    macd_warn = "⛽空中加油" if (curr["DIF"]>curr["DEA"] and curr["DIF"]>0 and curr["MACD_Bar"]>prev["MACD_Bar"]) else ""
    bar_trend = "🔴红增" if curr["MACD_Bar"] > 0 else "🟢绿缩"
    final_macd = f"{bar_trend}|{macd_warn if macd_warn else cross_status}"
    bb_state = "🚀突破上轨" if curr["BB_PctB"] > 1.0 else ("↔️极度收口" if curr["BB_Width"] < 12 else "")

    return {
        "代码": code, "名称": name, "现价": curr["close"],
        "今日涨跌": f"{today_pct:+.2f}%", "3日涨跌": f"{pct_3day:+.2f}%",
        "K线形态": kline_status, "K线评分": kline_score,
        "60分状态": status_60m, "BIAS乖离": round(curr["BIAS20"], 1),
        "连续": "", "共振因子": resonance_str,
        "信号类型": signal_type, "热门概念": display_concept,
        "OBV状态": "🚀健康流入",
        "筹码分布": chip_signal, "形态特征": pattern_str,
        "MACD状态": final_macd, "布林状态": bb_state,
        "今日CMF": round(curr["CMF"], 3), "昨日CMF": round(prev["CMF"], 3), "前日CMF": round(prev_2["CMF"], 3),
        "RSI指标": round(curr["RSI"], 1), "J值": round(curr["J"], 1),
        "建议挂单": suggest_buy, "止损价": stop_loss
    }

# --- 评分与排序 ---
def calculate_total_score(row):
    score = 0
    score += float(row.get('K线评分', 0)) # K线健康度
    
    s60 = str(row.get('60分状态', ''))
    if "金叉" in s60: score += 100    
    elif "多头" in s60: score += 80   
    elif "回调" in s60: score += 20   
    
    streak = str(row.get('连续', ''))
    if "3连" in streak or "4连" in streak: score += 50
    elif "2连" in streak: score += 30
    else: score += 10 
    
    try:
        c1, c2, c3 = float(row.get('今日CMF', 0)), float(row.get('昨日CMF', 0)), float(row.get('前日CMF', 0))
        if c1 > c2 > c3: score += 30 
        elif c1 > c2: score += 10
    except: pass
    
    if "黄金坑" in str(row.get('信号类型', '')): score += 20
    if "双金叉" in str(row.get('金叉信号', '')): score += 15
    if "筹码密集" in str(row.get('筹码分布', '')): score += 15
    if "🔥" in str(row.get('热门概念', '')): score += 10
    
    return score

# --- 历史与输出 ---
def update_history(current_results):
    today_str = datetime.now().strftime("%Y-%m-%d")
    try:
        if os.path.exists(HISTORY_FILE):
            hist_df = pd.read_csv(HISTORY_FILE)
            hist_df['date'] = hist_df['date'].astype(str)
        else: hist_df = pd.DataFrame(columns=["date", "code"])
    except: hist_df = pd.DataFrame(columns=["date", "code"])

    hist_df = hist_df[hist_df['date'] != today_str]
    sorted_dates = sorted(hist_df['date'].unique(), reverse=True)
    processed_results = []
    new_rows = []
    
    for res in current_results:
        code = res['code'] if 'code' in res else res['代码']
        streak = 1
        for d in sorted_dates:
            if not hist_df[(hist_df['date'] == d) & (hist_df['code'] == str(code))].empty: streak += 1
            else: break
        res['连续'] = f"🔥{streak}连" if streak >= 2 else "首榜"
        processed_results.append(res)
        new_rows.append({"date": today_str, "code": str(code)})

    if new_rows: hist_df = pd.concat([hist_df, pd.DataFrame(new_rows)], ignore_index=True)
    try: hist_df.to_csv(HISTORY_FILE, index=False)
    except: pass
    return processed_results

def save_and_beautify(data_list):
    dt_str = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"严选_作战地图版_{dt_str}.xlsx"
    
    if not data_list:
        pd.DataFrame([["无股入选 (条件严苛)"]]).to_excel(filename)
        return filename

    df = pd.DataFrame(data_list)
    df["综合评分"] = df.apply(calculate_total_score, axis=1)
    
    cols = ["代码", "名称", "综合评分", "现价", "今日涨跌", "3日涨跌", "K线形态", "60分状态", 
            "BIAS乖离", "连续", "共振因子", "信号类型", "热门概念", "OBV状态", "今日CMF", 
            "昨日CMF", "前日CMF", "筹码分布", "形态特征", "MACD状态", "布林状态", 
            "RSI指标", "J值", "建议挂单", "止损价"]
    for c in cols:
        if c not in df.columns: df[c] = ""
    df = df[cols]
    df.sort_values(by=["综合评分"], ascending=False, inplace=True)
    df.to_excel(filename, index=False)
    
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # 样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color="FFFFFF")
    fill_blue = PatternFill("solid", fgColor="4472C4")
    font_red = Font(color="FF0000", bold=True)
    font_green = Font(color="008000", bold=True)
    font_purple = Font(color="800080", bold=True)
    fill_yellow = PatternFill("solid", fgColor="FFF2CC")
    
    for cell in ws[1]:
        cell.fill = fill_blue
        cell.font = header_font
    
    for row in ws.iter_rows(min_row=2):
        if float(row[2].value) >= 150: row[2].fill = PatternFill("solid", fgColor="FFC7CE") 
        for idx in [4, 5]: 
            val = str(row[idx].value)
            if "+" in val: row[idx].font = font_red
            elif "-" in val: row[idx].font = font_green
        
        # K线形态
        k_val = str(row[6].value)
        if "强攻" in k_val or "仙人" in k_val: row[6].font = font_red
        elif "护盘" in k_val: row[6].font = font_purple
        elif "抛压" in k_val: row[6].font = font_green; row[6].fill = fill_yellow

        if "金叉" in str(row[7].value): row[7].font = font_red; row[7].fill = fill_yellow
        elif "回调" in str(row[7].value): row[7].font = font_green

        bias_val = row[8].value
        if isinstance(bias_val, (int, float)):
            if bias_val < -8: row[8].font = font_green; row[8].fill = fill_yellow
            elif bias_val > 12: row[8].font = font_red

        if "连" in str(row[9].value): row[9].font = font_red; row[9].fill = fill_yellow
        if "流入" in str(row[13].value): row[13].font = font_red
        if "红增" in str(row[19].value): row[19].font = font_red
        
        try:
            c1, c2, c3 = float(row[14].value), float(row[15].value), float(row[16].value)
            row[14].font = font_red
            if c1 > c2 > c3:
                row[14].fill = fill_yellow; row[15].font = font_red; row[16].font = font_red
        except: pass

        if "蚂蚁" in str(row[18].value): row[18].font = font_purple
        if "红肥" in str(row[18].value): row[18].font = font_red

    ws.column_dimensions['G'].width = 15 
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['K'].width = 25
    
    # ==========================================
    # 📚 终极作战地图 (The Combat Map)
    # ==========================================
    start_row = ws.max_row + 3
    title_font = Font(name='微软雅黑', size=14, bold=True, color="FFFFFF")
    cat_font = Font(name='微软雅黑', size=12, bold=True, color="0000FF")
    text_font = Font(name='微软雅黑', size=10)
    
    # --- 1. 五大策略实战手册 ---
    ws.cell(row=start_row, column=1, value="⚔️ 五大策略实战手册 (Strategy Manual)").font = cat_font
    start_row += 1
    strategies = [
        ("⚱️ 黄金坑", "【核心逻辑】深跌(BIAS<-8)后，今日放量阳线站稳MA5。左侧反转第一天。", "【买卖点】现价买入。止损设在前日最低点。"),
        ("🐉 龙回头", "【核心逻辑】前期妖股回调至生命线(MA60/MA20)附近，极致缩量。", "【买卖点】在'建议挂单'价位低吸。跌破布林下轨止损。"),
        ("🏦 机构控盘", "【核心逻辑】CMF>0.1(强吸筹) + ADX趋势向上 + 均线多头。", "【买卖点】沿5日线/10日线持股。"),
        ("📉 极度超跌", "【核心逻辑】RSI<20 或 底背离，且资金未流出。", "【买卖点】左侧分批买入，反弹5-10%即止盈。"),
        ("⚡ 底部变盘", "【核心逻辑】布林带宽<12(极度收口) + 资金异动。", "【买卖点】放量突破布林上轨瞬间追击。")
    ]
    for name, logic, action in strategies:
        ws.cell(row=start_row, column=1, value=name).font = Font(bold=True)
        ws.cell(row=start_row, column=2, value=logic).font = text_font
        ws.cell(row=start_row, column=3, value=action).font = text_font
        ws.merge_cells(start_row=start_row, start_column=3, end_row=start_row, end_column=10)
        start_row += 1
    start_row += 1
    
    # --- 2. 全指标读图指南 ---
    ws.cell(row=start_row, column=1, value="📊 全指标读图指南 (Reading Guide)").font = cat_font
    start_row += 1
    indicators = [
        ("K线形态", "💪实体强攻：多头强势(最好)；🛡️下影护盘：主力托底(安全)；☝️仙人指路：上涨中继(加仓)；⚠️抛压沉重：高位风险(减仓)。"),
        ("60分状态", "✅金叉(黄底)：日内最佳买点；🚀多头(红字)：持股/顺势买；⚠️回调(绿字)：日线好但短线跌，建议等金叉再买。"),
        ("CMF三日", "资金流向指标。若[前<昨<今]且标黄，代表主力不计成本加速抢筹，爆发力最强。"),
        ("BIAS乖离", "<-8% (绿黄底)：黄金坑区域，机会大； >12% (红字)：短线超买，谨防回调。"),
        ("MACD状态", "🔴红增：多头增强；🟢绿缩：空头衰竭；⛽空中加油：上涨中继(强)。"),
        ("形态特征", "🟥红肥绿瘦：倍量吸筹；🐜蚂蚁上树：温和建仓；⚡N字反包：强势洗盘。"),
        ("共振因子", "显示该股满足的核心条件(如 策略+热点+双金叉)。满足越多，确定性越高。"),
        ("止损价", "⛔ 风控铁律！收盘价跌破此价格，说明逻辑破坏，必须无条件卖出。")
    ]
    for name, desc in indicators:
        ws.cell(row=start_row, column=1, value=name).font = Font(bold=True)
        ws.cell(row=start_row, column=2, value=desc).font = text_font
        ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=10)
        start_row += 1

    wb.save(filename)
    print(f"✅ 结果已保存: {filename}")
    return filename

def analyze_one_stock(code, name, start_dt):
    try:
        df = get_data_with_retry(code, start_dt)
        if df is None: return None
        return process_stock_logic(df, code, name)
    except: return None

def main():
    print("=== A股严选 (全攻略·作战地图版) ===")
    get_market_hot_spots()
    start_time = time.time()
    targets, source_name = get_targets_robust()
    start_dt = (datetime.now() - timedelta(days=CONFIG["DAYS_LOOKBACK"])).strftime("%Y%m%d")
    
    print(f"[{source_name}] 待扫描: {len(targets)} 只 | 启动 {CONFIG['MAX_WORKERS']} 线程...")
    results = []
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=CONFIG["MAX_WORKERS"]) as executor:
        future_to_stock = {executor.submit(analyze_one_stock, r['code'], r['name'], start_dt): r['code'] for _, r in targets.iterrows()}
        count = 0
        total = len(targets)
        for future in concurrent.futures.as_completed(future_to_stock):
            count += 1
            if count % 100 == 0: print(f"进度: {count}/{total} ...")
            try:
                res = future.result()
                if res:
                    print(f"  ★ 严选: {res['名称']} [{res['信号类型']}] BIAS:{res['BIAS乖离']}")
                    results.append(res)
            except: pass

    if results: results = update_history(results)
    print(f"\n耗时: {int(time.time() - start_time)}秒 | 选中 {len(results)} 只")
    save_and_beautify(results)

if __name__ == "__main__":
    main()
